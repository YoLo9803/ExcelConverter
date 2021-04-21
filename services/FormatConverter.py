from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import numbers

class FormatConverter():
    def __init__(self):
        self.__data = None
        self.__columnMap = {1: 0, 2: 2, 3: 9, 4: 15, 5: 18, 6: 20, 7: 22, 8: 25}

    def process(self, workbook):
        result = Workbook()
        ws = result.active
        sheet = workbook.get_sheet_by_name('廣代')
        
        for index, row in enumerate(sheet.rows):
            text = "收件人：" + row[0].value + "\n電話：" + row[2].value + "\n地址：" + row[1].value 
            print(text)
            ws.cell(column = 1, row = index + 1).alignment = Alignment(wrapText=True)
            ws.cell(column = 1, row = index + 1).value = text
        return result